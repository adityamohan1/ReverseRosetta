"""End-to-end smoke tests with deterministic codon generator (no HuggingFace)."""

from __future__ import annotations

from pathlib import Path

import pytest

from reverserosetta.config import ReverseRosettaConfig
from reverserosetta.excel_io import read_amino_acid_column
from reverserosetta.optimize import optimize_sequences_from_dataframe
from reverserosetta.splice import HeuristicSpliceBackend
from reverserosetta.utils import AA_TO_CODONS, HUMAN_CODON_WEIGHT


def _preferred_dna(aa: str, **_kwargs: object) -> str:
    parts: list[str] = []
    for a in aa:
        opts = AA_TO_CODONS[a]
        best = max(opts, key=lambda c: HUMAN_CODON_WEIGHT.get(c, 0.0))
        parts.append(best)
    return "".join(parts)


def test_smoke_optimize_from_temp_excel(tmp_path: Path) -> None:
    import pandas as pd

    xlsx = tmp_path / "in.xlsx"
    # Short sequence keeps the smoke test fast and avoids rare motif clashes.
    pd.DataFrame([["id", "n", "x", "MG"]]).to_excel(
        xlsx, sheet_name="Sheet1", index=False, header=False
    )
    df = read_amino_acid_column(xlsx, sheet_name="Sheet1", column_index=4)
    cfg = ReverseRosettaConfig(max_iterations=200, emit_stop_codon=False)
    be = HeuristicSpliceBackend()
    out = optimize_sequences_from_dataframe(df, cfg, splice_backend=be, codon_fn=_preferred_dna)
    assert len(out) >= 1
    for row in out:
        assert len(row.final_dna) == 3 * len(row.aa_sequence)
        assert row.validation_ok


def test_empty_excel_column_raises_or_skips(tmp_path: Path) -> None:
    from openpyxl import Workbook

    p = tmp_path / "empty.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for r, a, b, c in ((1, 1, 2, 3), (2, 4, 5, 6)):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        ws.cell(row=r, column=3, value=c)
        ws.cell(row=r, column=4, value="")
    wb.save(p)
    df2 = read_amino_acid_column(p, "S", 4)
    assert df2.empty


def test_invalid_aa_rejected(tmp_path: Path) -> None:
    import pandas as pd

    from reverserosetta.excel_io import InvalidAminoAcidError

    p = tmp_path / "bad.xlsx"
    df = pd.DataFrame(
        [
            [1, 2, 3, "MK"],
            [1, 2, 3, "M1"],
        ]
    )
    with pd.ExcelWriter(p, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="S", index=False, header=False)
    with pytest.raises(InvalidAminoAcidError):
        read_amino_acid_column(p, "S", 4)
